from openpyxl import load_workbook, Workbook, styles
import requests
import time
from bs4 import BeautifulSoup

wb = load_workbook('practice_names.xlsx')
wb2 = Workbook()
out = wb2.active
out["A1"] = "Name"
out['A1'].font = styles.Font(color=styles.colors.BLUE)  # http://bit.ly/2jAeYM5
out["B1"] = "Address"
out['B1'].font = styles.Font(color=styles.colors.BLUE)
dest_filename = 'output.xlsx'  # print wb.get_sheet_names()
sheet = wb['Engagement  Total']
cell = "A%s"  # names

# https://www.healthgrades.com/api3/autosuggest/what?term=HERNANDEZ,%20ANDRES&pt=40.710099,-73.948845
api = 'https://www.healthgrades.com/api3/autosuggest/what'
point = "40.710099, -73.948845"  # brooklyn NY
params = dict(pt=point, term='')


row = 2
while 1:
    current = cell % row
    val = sheet[current].value
    if not val:
        break

    name = val.upper()
    params["term"] = name
    response = requests.get(api, params=params)
    json = response.json()
    print json
    categories = json["response"]["categories"]
    not_found = False
    if not categories:
        not_found = True
    else:
        providers = None

        for each in categories:
            if each['title'] == "Healthcare Providers Near {Location}":
                providers = each["suggestions"]
            else:
                not_found = True

        new_york = None
        if providers:
            new_york = filter(lambda e: ", NY" in e["highlightedText"], providers)

        if new_york:
            start_column = ord("B")
            for i, each in enumerate(new_york):
                next_cell = chr(start_column + i) + "%s"  # B, C, D...
                profile_url = each["entity"]["profileUrl"]
                if profile_url:
                    html = requests.get("https://www.healthgrades.com" + profile_url).content
                    soup = BeautifulSoup(html)
                    out["B%s" % row] = soup.find("address").get_text(" ")  # join text with space http://bit.ly/2jAsRtF
                else:
                    out["B%s" % row] = "No Healthgrades Profile"
                    out["B%s" % row].font = styles.Font(color=styles.colors.RED)
        else:
            out["B%s" % row] = "Name Not Found In NY"
            out["B%s" % row].font = styles.Font(color=styles.colors.RED)

    if not_found:
        out["B%s" % row] = "Provider Not Found"
        out["B%s" % row].font = styles.Font(color=styles.colors.RED)
    out[current] = name

    time.sleep(0.15)
    row += 1

wb2.save(filename=dest_filename)
